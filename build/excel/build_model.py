# -*- coding: utf-8 -*-
"""
Сборка Excel-модели pay-gap калькулятора (шаг 3 по SPEC §8).

Читает CSV той же формы, что и калькулятор, и строит .xlsx, в котором вся
денежная цепочка живая: нормализация -> разрывы -> регрессия (LINEST) ->
отбор адресатов -> распределение прибавки -> взносы -> итоги.

Из Python приходят только исходные данные. Ни одна сумма не вписывается
числом: всё, что можно пересчитать, пересчитывается формулой Excel.

Соответствие calc.js:
  normalise()        -> лист Data, колонки J..N
  regress()          -> лист Categories, LINEST + guard-условия
  computeGaps()      -> лист Categories
  runScenario()      -> лист Data (адресаты) + лист Categories (суммы)
  contributionsOn()  -> лист Data, колонки взносов
  totalsBlock()      -> лист Summary
"""
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# --- Настройки по умолчанию: DEFAULT_SETTINGS в calc.js ---------------------
DEFAULTS = {
    "threshold_pct": 5,
    "implementation_month": 7,
    "rate_below": 31.5,
    "rate_above": 1.15,
    "ceiling": 61214,
}

# --- Оформление ------------------------------------------------------------
H_FILL = PatternFill("solid", fgColor="1F3B57")
H_FONT = Font(color="FFFFFF", bold=True, size=10)
IN_FILL = PatternFill("solid", fgColor="FFF4D6")    # исходные данные
CALC_FILL = PatternFill("solid", fgColor="EAF3FA")  # расчёт формулой
SET_FILL = PatternFill("solid", fgColor="E6F4EA")   # настройка, можно менять
TITLE = Font(bold=True, size=13)
BOLD = Font(bold=True)
MUTED = Font(color="6B7280", size=9)
THIN = Side(style="thin", color="D0D7DE")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EUR = '#,##0.00'
PCT = '0.00"%"'
NUM = '#,##0.0000'


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    out = []
    for i, r in enumerate(rows):
        out.append({
            "id": r.get("id") or "E%03d" % (i + 1),
            "category": r["category"].strip(),
            "gender": "F" if r["gender"].strip().upper() == "F" else "M",
            "base_salary": float(r["base_salary"]),
            "variable_pay": float(r["variable_pay"]),
            "tenure_years": float(r["tenure_years"]),
            "grade": float(r["grade"]),
            "fte": float(r["fte"]) if r.get("fte") else 1.0,
            "months_worked": float(r["months_worked"]) if r.get("months_worked") else 12.0,
        })
    return out


def header(ws, row, labels, widths=None):
    for j, text in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=text)
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 30


def build(csv_path, out_path, title_note):
    emp = read_csv(csv_path)
    n = len(emp)

    # порядок категорий — первого появления, как в analyse()
    order, seen = [], set()
    for e in emp:
        if e["category"] not in seen:
            seen.add(e["category"])
            order.append(e["category"])

    wb = Workbook()

    # ======================================================================
    # ЛИСТ Settings
    # ======================================================================
    st = wb.active
    st.title = "Settings"
    st["A1"] = "Model settings"
    st["A1"].font = TITLE
    st["A2"] = title_note
    st["A2"].font = MUTED
    st["A4"] = "Only the green cells are meant to be edited. Everything else recalculates."
    st["A4"].font = MUTED

    rows_set = [
        ("threshold_pct", DEFAULTS["threshold_pct"], "Threshold, %",
         "Target unexplained gap for the Minimum compliance scenario."),
        ("implementation_month", DEFAULTS["implementation_month"], "Implementation month",
         "1 = January. Drives the current-year effect: (13 - m) / 12."),
        ("rate_below", DEFAULTS["rate_below"], "Contribution rate below ceiling, %",
         "Applied to the part of the uplift that fits under the ceiling."),
        ("rate_above", DEFAULTS["rate_above"], "Contribution rate above ceiling, %",
         "Applied to the part that pushes base pay past the ceiling."),
        ("ceiling", DEFAULTS["ceiling"], "Contribution ceiling, EUR",
         "Annual, applied to the employee's actual base pay."),
    ]
    header(st, 6, ["Name", "Value", "What it is", "Notes"], [24, 14, 34, 76])
    for i, (name, val, label, note) in enumerate(rows_set):
        r = 7 + i
        st.cell(row=r, column=1, value=name).border = BOX
        c = st.cell(row=r, column=2, value=val)
        c.fill = SET_FILL
        c.border = BOX
        c.font = BOLD
        st.cell(row=r, column=3, value=label).border = BOX
        d = st.cell(row=r, column=4, value=note)
        d.border = BOX
        d.alignment = Alignment(wrap_text=True, vertical="top")
        wb.defined_names.add(DefinedName(name, attr_text="Settings!$B$%d" % r))
        st.row_dimensions[r].height = 28

    st["A14"] = "Source data"
    st["A14"].font = BOLD
    st["A15"] = ("The Data sheet holds a fixed export. The demo-data generator uses a PRNG "
                 "(mulberry32 + Box-Muller) that cannot be reproduced byte-for-byte in Excel "
                 "formulas, so the array is exported as plain numbers. Everything computed "
                 "on top of it is live.")
    st["A15"].font = MUTED
    st["A15"].alignment = Alignment(wrap_text=True, vertical="top")
    st.merge_cells("A15:D17")

    # ======================================================================
    # ЛИСТ Data
    # ======================================================================
    ws = wb.create_sheet("Data")
    ws["A1"] = "Employee-level data and calculation"
    ws["A1"].font = TITLE
    ws["A2"] = ("Yellow — source data from the CSV. Blue — calculated by formula. "
                "Formulas reference the Categories sheet, where category-level gaps are computed.")
    ws["A2"].font = MUTED

    cols = [
        "id", "category", "gender",
        "base_salary", "variable_pay",
        "tenure_years", "grade", "fte", "months_worked",
        "factor", "n_base", "n_var", "n_total",
        "actual_total",
        "median_cat", "recip_min", "deficit_min", "share_min",
        "uplift_norm_min", "uplift_act_min", "contrib_min", "cost_min",
        "recip_full", "deficit_full", "share_full",
        "uplift_norm_full", "uplift_act_full", "contrib_full", "cost_full",
        "_rank_cat", "_n_cat", "_rank_sex", "_n_sex",
    ]
    HDR = 4
    header(ws, HDR, cols, [9, 18, 8, 13, 13, 11, 8, 7, 11,
                           9, 12, 12, 12, 13,
                           12, 11, 12, 10, 13, 13, 12, 12,
                           11, 12, 10, 13, 13, 12, 12,
                           10, 8, 10, 8])
    ws.freeze_panes = "A%d" % (HDR + 1)

    # единицы измерения — примечаниями к заголовкам, а не в самих названиях:
    # длинные подписи в шапке загромождают таблицу
    notes = {
        4: "EUR, actually paid for the period. Not annualised.",
        5: "EUR, actually paid for the period. Recorded on the same basis as base "
           "salary: what was actually paid, not a full-year figure.",
        7: "Position in the grade ladder: higher number = higher grade. Only the "
           "ordering matters for the calculation, not the absolute value.",
        8: "Share of a full-time position: 1.0 = full time, 0.5 = half time.",
        9: "Months worked during the period. 12 = full year.",
        10: "factor = fte x (months_worked / 12). Converts actual pay to a "
            "full-time, full-year equivalent.",
        11: "EUR per year at full time: base_salary / factor.",
        12: "EUR per year at full time: variable_pay / factor.",
        13: "EUR per year at full time: n_base + n_var. This is the figure "
            "employees are compared on.",
        14: "EUR, actually paid: base_salary + variable_pay. The payroll total "
            "is the sum of this column.",
        15: "Median n_total across the whole category, both genders. "
            "EUR per year at full time.",
    }
    for col, text in notes.items():
        ws.cell(row=HDR, column=col).comment = Comment(text, "Model")

    first, last = HDR + 1, HDR + n

    # строки категорий на листе Categories (заголовок там в строке CHDR_CAT)
    CHDR_CAT = 4
    CFIRST = CHDR_CAT + 1
    CLAST = CFIRST + len(order) - 1

    def cat_lookup(col, row):
        """Значение из колонки col листа Categories для категории строки row.

        Поиск по имени категории, а не по номеру строки: SUMIFS возвращает
        единственное совпадение, потому что имена категорий уникальны.
        """
        return ("SUMIFS(Categories!${c}${f}:${c}${l},Categories!$A${f}:$A${l},B{r})"
                .format(c=col, f=CFIRST, l=CLAST, r=row))

    for i, e in enumerate(emp):
        r = first + i

        vals = [e["id"], e["category"], e["gender"], e["base_salary"],
                e["variable_pay"], e["tenure_years"], e["grade"],
                e["fte"], e["months_worked"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.fill = IN_FILL
            c.border = BOX
            if j in (4, 5):
                c.number_format = EUR

        def put(col, formula, fmt=None):
            c = ws.cell(row=r, column=col, value=formula)
            c.fill = CALC_FILL
            c.border = BOX
            if fmt:
                c.number_format = fmt
            return c

        # --- нормализация: normalise() ---
        put(10, "=H{r}*(I{r}/12)".format(r=r), NUM)      # factor
        put(11, "=D{r}/J{r}".format(r=r), EUR)           # n_base
        put(12, "=E{r}/J{r}".format(r=r), EUR)           # n_var
        put(13, "=K{r}+L{r}".format(r=r), EUR)           # n_total
        put(14, "=D{r}+E{r}".format(r=r), EUR)           # actual_total

        # медиана категории (оба пола), без формул массива.
        # SUMPRODUCT считает ранг строки внутри её категории по n_total, а
        # медиана собирается из одного (нечётное n) или двух (чётное) средних
        # значений через SUMIFS по этому рангу. Обычные формулы, Ctrl+Shift+Enter
        # не нужен ни в одной версии Excel.
        rank = ("SUMPRODUCT(($B${f}:$B${l}=B{r})*($M${f}:$M${l}<M{r}))+1"
                .format(f=first, l=last, r=r))
        put(15, ("=(SUMIFS($M${f}:$M${l},$B${f}:$B${l},B{r},$AD${f}:$AD${l},"
                 "INT((AE{r}+1)/2))+SUMIFS($M${f}:$M${l},$B${f}:$B${l},B{r},"
                 "$AD${f}:$AD${l},INT(AE{r}/2)+1))/2").format(
            f=first, l=last, r=r), EUR)

        # --- сценарий minimum ---
        # Значения с листа Categories берутся поиском по имени категории
        # (SUMIFS по колонке A), а не ссылкой на конкретную строку: строка
        # категории может сдвинуться, имя — нет.
        put(16, '=IF(AND(C{r}="F",M{r}<O{r},{corr}=1),1,0)'.format(
            r=r, corr=cat_lookup("Z", r)))
        put(17, "=IF(P{r}=1,O{r}-M{r},0)".format(r=r), EUR)
        put(18, "=IF(AND(P{r}=1,{td}>0),Q{r}/{td},0)".format(
            r=r, td=cat_lookup("AA", r)), NUM)
        put(19, "=R{r}*{need}".format(r=r, need=cat_lookup("AB", r)), EUR)
        put(20, "=S{r}*J{r}".format(r=r), EUR)
        put(21, ("=IF(T{r}<=0,0,"
                 "(rate_below/100)*MIN(T{r},MAX(0,ceiling-D{r}))"
                 "+(rate_above/100)*(T{r}-MIN(T{r},MAX(0,ceiling-D{r}))))").format(r=r), EUR)
        put(22, "=T{r}+U{r}".format(r=r), EUR)

        # --- сценарий full ---
        put(23, '=IF(AND(C{r}="F",M{r}<O{r},{corr}=1),1,0)'.format(
            r=r, corr=cat_lookup("AH", r)))
        put(24, "=IF(W{r}=1,O{r}-M{r},0)".format(r=r), EUR)
        put(25, "=IF(AND(W{r}=1,{td}>0),X{r}/{td},0)".format(
            r=r, td=cat_lookup("AI", r)), NUM)
        put(26, "=Y{r}*{need}".format(r=r, need=cat_lookup("AJ", r)), EUR)
        put(27, "=Z{r}*J{r}".format(r=r), EUR)
        put(28, ("=IF(AA{r}<=0,0,"
                 "(rate_below/100)*MIN(AA{r},MAX(0,ceiling-D{r}))"
                 "+(rate_above/100)*(AA{r}-MIN(AA{r},MAX(0,ceiling-D{r}))))").format(r=r), EUR)
        put(29, "=AA{r}+AB{r}".format(r=r), EUR)

        # --- служебные колонки для немассивной медианы ---
        # AD: ранг строки по n_total внутри своей категории (1 = самый низкий).
        # Второе слагаемое разрывает ничьи по номеру строки: при равных n_total
        # ранги всё равно различаются, иначе SUMIFS по рангу вернул бы сумму
        # двух строк вместо одной и медиана оказалась бы вдвое больше.
        put(30, ("=SUMPRODUCT(($B${f}:$B${l}=B{r})*($M${f}:$M${l}<M{r}))"
                 "+SUMPRODUCT(($B${f}:$B${l}=B{r})*($M${f}:$M${l}=M{r})"
                 "*(ROW($M${f}:$M${l})<ROW(M{r})))+1"
                 ).format(f=first, l=last, r=r))
        # AE: размер категории
        put(31, "=COUNTIF($B${f}:$B${l},B{r})".format(f=first, l=last, r=r))
        # AF: ранг внутри пары «категория + пол» — для медиан по полу
        put(32, ("=SUMPRODUCT(($B${f}:$B${l}=B{r})*($C${f}:$C${l}=C{r})"
                 "*($M${f}:$M${l}<M{r}))"
                 "+SUMPRODUCT(($B${f}:$B${l}=B{r})*($C${f}:$C${l}=C{r})"
                 "*($M${f}:$M${l}=M{r})*(ROW($M${f}:$M${l})<ROW(M{r})))+1"
                 ).format(f=first, l=last, r=r))
        # AG: размер группы «категория + пол»
        put(33, "=COUNTIFS($B${f}:$B${l},B{r},$C${f}:$C${l},C{r})".format(
            f=first, l=last, r=r))

    # ======================================================================
    # ЛИСТ Categories
    # ======================================================================
    cs = wb.create_sheet("Categories")
    cs["A1"] = "Category-level calculation"
    cs["A1"].font = TITLE
    cs["A2"] = ("Regression: OLS of n_total on grade and tenure within each category, "
                "solved from explicit sums (no LINEST, no array formulas). Guard conditions "
                "mirror calc.js: fewer than 4 observations or a constant predictor means the "
                "regression is undefined, and the whole gap is treated as unexplained.")
    cs["A2"].font = MUTED

    ccols = [
        "category", "headcount", "n_F", "n_M",                          # A-D
        "mean_M", "mean_F", "median_all",                               # E-G
        "raw_gap_mean_%", "median_M", "median_F", "raw_gap_median_%",   # H-K
        "var_M", "var_F", "variable_gap_%",                             # L-N
        "base_M", "base_F", "base_gap_%",                               # O-Q
        "var_grade", "var_tenure", "reg_ok",                            # R-T
        "coef_grade", "coef_tenure", "r2",                              # U-W
        "explained_%", "unexplained_%",                                 # X-Y
        "correct_min", "totalDef_min", "need_min",                      # Z-AB
        "adjust_min", "contrib_min", "cost_min", "recip_min",           # AC-AF
        "unexpl_after_min",                                             # AG
        "correct_full", "totalDef_full", "need_full",                   # AH-AJ
        "adjust_full", "contrib_full", "cost_full", "recip_full",       # AK-AN
        "unexpl_after_full",                                            # AO
        "unreliable", "reverse_gap",                                    # AP-AQ
    ] + [""] * 9 + [                                                    # AR-AZ пусто
        "n", "Sum_g", "Sum_t", "Sum_y", "Sum_gg", "Sum_tt",             # BA-BF
        "Sum_gt", "Sum_gy", "Sum_ty", "Sum_yy", "det",                  # BG-BK
    ]
    CHDR = CHDR_CAT   # одна и та же величина: Data ссылается на эту разметку
    header(cs, CHDR, ccols, [18, 10, 7, 7] + [12] * (len(ccols) - 4))
    cs.freeze_panes = "B%d" % (CHDR + 1)

    D = "Data!"

    def rng(col):
        return "%s$%s$%d:$%s$%d" % (D, col, first, col, last)

    B, G_ = rng("B"), rng("C")          # категория, пол
    M_, K_, L_ = rng("M"), rng("K"), rng("L")   # n_total, n_base, n_var
    GR, TE = rng("G"), rng("F")         # grade, tenure

    for i, name in enumerate(order):
        r = CHDR + 1 + i
        cf = "$A%d" % r

        def put(col, formula, fmt=None, fill=CALC_FILL):
            c = cs.cell(row=r, column=col, value=formula)
            c.fill = fill
            c.border = BOX
            if fmt:
                c.number_format = fmt
            return c

        cat = cs.cell(row=r, column=1, value=name)
        cat.fill = IN_FILL
        cat.border = BOX

        put(2, "=COUNTIF({b},{c})".format(b=B, c=cf))
        put(3, '=COUNTIFS({b},{c},{g},"F")'.format(b=B, c=cf, g=G_))
        put(4, '=COUNTIFS({b},{c},{g},"M")'.format(b=B, c=cf, g=G_))

        put(5, '=IF(D{r}=0,0,AVERAGEIFS({m},{b},{c},{g},"M"))'.format(
            r=r, m=M_, b=B, c=cf, g=G_), EUR)
        put(6, '=IF(C{r}=0,0,AVERAGEIFS({m},{b},{c},{g},"F"))'.format(
            r=r, m=M_, b=B, c=cf, g=G_), EUR)
        # медиана всей категории — берём готовую из Data (там уже посчитана
        # немассивной формулой, у всех строк категории она одинаковая)
        put(7, "=IF(B{r}=0,0,SUMIFS({o},{b},{c},{ad},1))".format(
            r=r, o=rng("O"), b=B, c=cf, ad=rng("AD")), EUR)
        put(8, "=IF(OR(E{r}<=0,C{r}=0,D{r}=0),0,(E{r}-F{r})/E{r}*100)".format(r=r), PCT)

        # медианы по полу — через ранг внутри группы «категория + пол» (Data!AF).
        # При чётном размере группы берутся два средних значения, при нечётном
        # обе половины формулы указывают на одну и ту же строку.
        def med_by_sex(sex, cnt_cell):
            return ('=IF({cnt}=0,0,(SUMIFS({m},{b},{c},{g},"{s}",{af},'
                    'INT(({cnt}+1)/2))+SUMIFS({m},{b},{c},{g},"{s}",{af},'
                    'INT({cnt}/2)+1))/2)').format(
                m=M_, b=B, c=cf, g=G_, s=sex, af=rng("AF"), cnt=cnt_cell)

        put(9, med_by_sex("M", "D%d" % r), EUR)
        put(10, med_by_sex("F", "C%d" % r), EUR)
        put(11, "=IF(OR(I{r}<=0,C{r}=0,D{r}=0),0,(I{r}-J{r})/I{r}*100)".format(r=r), PCT)

        put(12, '=IF(D{r}=0,0,AVERAGEIFS({l},{b},{c},{g},"M"))'.format(
            r=r, l=L_, b=B, c=cf, g=G_), EUR)
        put(13, '=IF(C{r}=0,0,AVERAGEIFS({l},{b},{c},{g},"F"))'.format(
            r=r, l=L_, b=B, c=cf, g=G_), EUR)
        put(14, "=IF(OR(L{r}<=0,C{r}=0,D{r}=0),0,(L{r}-M{r})/L{r}*100)".format(r=r), PCT)

        put(15, '=IF(D{r}=0,0,AVERAGEIFS({k},{b},{c},{g},"M"))'.format(
            r=r, k=K_, b=B, c=cf, g=G_), EUR)
        put(16, '=IF(C{r}=0,0,AVERAGEIFS({k},{b},{c},{g},"F"))'.format(
            r=r, k=K_, b=B, c=cf, g=G_), EUR)
        put(17, "=IF(OR(O{r}<=0,C{r}=0,D{r}=0),0,(O{r}-P{r})/O{r}*100)".format(r=r), PCT)

        # --- регрессия: guard как в regress() ---
        # дисперсия предиктора через Var = E[x^2] - (E[x])^2 — немассивно
        put(18, ("=IF(B{r}<2,0,SUMPRODUCT(({b}={c})*{g}*{g})/B{r}"
                 "-(SUMIFS({g},{b},{c})/B{r})^2)").format(
            r=r, b=B, c=cf, g=GR), NUM)
        put(19, ("=IF(B{r}<2,0,SUMPRODUCT(({b}={c})*{t}*{t})/B{r}"
                 "-(SUMIFS({t},{b},{c})/B{r})^2)").format(
            r=r, b=B, c=cf, t=TE), NUM)
        put(20, "=IF(AND(B{r}>=4,R{r}>0.000000001,S{r}>0.000000001),1,0)".format(r=r))

        # --- OLS без LINEST: нормальные уравнения 3x3, решение по Крамеру ---
        # Сначала суммы (колонки BA..BI ниже), затем центрированные моменты:
        #   Sgg = Σg² - (Σg)²/n     Sgt = Σgt - Σg·Σt/n
        #   Stt = Σt² - (Σt)²/n     Sgy = Σgy - Σg·Σy/n     Sty = Σty - Σt·Σy/n
        # Определитель системы 2x2 (после исключения свободного члена):
        #   det = Sgg·Stt - Sgt²
        #   coef_grade  = (Sgy·Stt - Sty·Sgt) / det
        #   coef_tenure = (Sty·Sgg - Sgy·Sgt) / det
        # Это ровно то, что solveGauss() считает методом Гаусса, но выписанное
        # явно — в 3x3 правило Крамера короче и целиком видно в ячейке.
        # служебные суммы, строго по колонкам BA(53)..BJ(62)
        sums = [
            ("BA", "=B{r}".format(r=r)),                                        # n
            ("BB", "=SUMIFS({g},{b},{c})".format(g=GR, b=B, c=cf)),             # Σg
            ("BC", "=SUMIFS({t},{b},{c})".format(t=TE, b=B, c=cf)),             # Σt
            ("BD", "=SUMIFS({m},{b},{c})".format(m=M_, b=B, c=cf)),             # Σy
            ("BE", "=SUMPRODUCT(({b}={c})*{g}*{g})".format(b=B, c=cf, g=GR)),   # Σgg
            ("BF", "=SUMPRODUCT(({b}={c})*{t}*{t})".format(b=B, c=cf, t=TE)),   # Σtt
            ("BG", "=SUMPRODUCT(({b}={c})*{g}*{t})".format(b=B, c=cf, g=GR, t=TE)),
            ("BH", "=SUMPRODUCT(({b}={c})*{g}*{m})".format(b=B, c=cf, g=GR, m=M_)),
            ("BI", "=SUMPRODUCT(({b}={c})*{t}*{m})".format(b=B, c=cf, t=TE, m=M_)),
            ("BJ", "=SUMPRODUCT(({b}={c})*{m}*{m})".format(b=B, c=cf, m=M_)),   # Σyy
        ]
        for k, (_col, formula) in enumerate(sums):
            put(53 + k, formula, NUM)
        Sgg = "(BE{r}-BB{r}^2/B{r})".format(r=r)
        Stt = "(BF{r}-BC{r}^2/B{r})".format(r=r)
        Sgt = "(BG{r}-BB{r}*BC{r}/B{r})".format(r=r)
        Sgy = "(BH{r}-BB{r}*BD{r}/B{r})".format(r=r)
        Sty = "(BI{r}-BC{r}*BD{r}/B{r})".format(r=r)
        Syy = "(BJ{r}-BD{r}^2/B{r})".format(r=r)
        det = "({gg}*{tt}-{gt}^2)".format(gg=Sgg, tt=Stt, gt=Sgt)

        # BK: определитель — вынесен отдельно, чтобы guard мог его проверить
        put(63, "=IF(T{r}=0,0,{det})".format(r=r, det=det), NUM)
        put(21, "=IF(OR(T{r}=0,ABS(BK{r})<0.000000001),0,({gy}*{tt}-{ty}*{gt})/BK{r})".format(
            r=r, gy=Sgy, tt=Stt, ty=Sty, gt=Sgt), EUR)
        put(22, "=IF(OR(T{r}=0,ABS(BK{r})<0.000000001),0,({ty}*{gg}-{gy}*{gt})/BK{r})".format(
            r=r, ty=Sty, gg=Sgg, gy=Sgy, gt=Sgt), EUR)
        # R² = (coef_grade·Sgy + coef_tenure·Sty) / Syy — доля объяснённой дисперсии
        put(23, ('=IF(OR(T{r}=0,ABS(BK{r})<0.000000001,{yy}<0.000000001),"",'
                 '(U{r}*{gy}+V{r}*{ty})/{yy})').format(
            r=r, yy=Syy, gy=Sgy, ty=Sty), NUM)

        # explained: coef_grade * dGrade + coef_tenure * dTenure, в % к mean_M
        dg = ('(IF(D{r}=0,0,AVERAGEIFS({g},{b},{c},{gd},"M"))'
              '-IF(C{r}=0,0,AVERAGEIFS({g},{b},{c},{gd},"F")))').format(
            r=r, g=GR, b=B, c=cf, gd=G_)
        dt = ('(IF(D{r}=0,0,AVERAGEIFS({t},{b},{c},{gd},"M"))'
              '-IF(C{r}=0,0,AVERAGEIFS({t},{b},{c},{gd},"F")))').format(
            r=r, t=TE, b=B, c=cf, gd=G_)
        raw_expl = "IF(OR(T{r}=0,E{r}<=0),0,(U{r}*{dg}+V{r}*{dt})/E{r}*100)".format(
            r=r, dg=dg, dt=dt)
        put(24, "=IF(H{r}>=0,MAX(0,MIN({e},H{r})),MIN(0,MAX({e},H{r})))".format(
            r=r, e=raw_expl), PCT)
        put(25, "=H{r}-X{r}".format(r=r), PCT)

        # --- сценарий minimum ---
        put(26, "=IF(AND(H{r}>=0,Y{r}>threshold_pct),1,0)".format(r=r))
        put(27, "=IF(Z{r}=0,0,SUMIFS({q},{b},{c}))".format(
            r=r, q="%s$Q$%d:$Q$%d" % (D, first, last), b=B, c=cf), EUR)
        put(28, "=IF(Z{r}=0,0,(Y{r}-threshold_pct)/100*E{r}*C{r})".format(r=r), EUR)
        put(29, "=SUMIFS({t},{b},{c})".format(
            r=r, t="%s$T$%d:$T$%d" % (D, first, last), b=B, c=cf), EUR)
        put(30, "=SUMIFS({u},{b},{c})".format(
            r=r, u="%s$U$%d:$U$%d" % (D, first, last), b=B, c=cf), EUR)
        put(31, "=AC{r}+AD{r}".format(r=r), EUR)
        put(32, "=SUMIFS({p},{b},{c})".format(
            r=r, p="%s$P$%d:$P$%d" % (D, first, last), b=B, c=cf))
        new_f = ('(IF(C{r}=0,0,AVERAGEIFS({m},{b},{c},{g},"F"))'
                 '+IF(C{r}=0,0,SUMIFS({s},{b},{c},{g},"F")/C{r}))').format(
            r=r, m=M_, b=B, c=cf, g=G_, s="%s$S$%d:$S$%d" % (D, first, last))
        put(33, "=IF(E{r}<=0,0,(E{r}-{nf})/E{r}*100-X{r})".format(r=r, nf=new_f), PCT)

        # --- сценарий full: цель 0 ---
        put(34, "=IF(AND(H{r}>=0,Y{r}>0),1,0)".format(r=r))
        put(35, "=IF(AH{r}=0,0,SUMIFS({x},{b},{c}))".format(
            r=r, x="%s$X$%d:$X$%d" % (D, first, last), b=B, c=cf), EUR)
        put(36, "=IF(AH{r}=0,0,Y{r}/100*E{r}*C{r})".format(r=r), EUR)
        put(37, "=SUMIFS({aa},{b},{c})".format(
            r=r, aa="%s$AA$%d:$AA$%d" % (D, first, last), b=B, c=cf), EUR)
        put(38, "=SUMIFS({ab},{b},{c})".format(
            r=r, ab="%s$AB$%d:$AB$%d" % (D, first, last), b=B, c=cf), EUR)
        put(39, "=AK{r}+AL{r}".format(r=r), EUR)
        put(40, "=SUMIFS({w},{b},{c})".format(
            r=r, w="%s$W$%d:$W$%d" % (D, first, last), b=B, c=cf))
        new_f2 = ('(IF(C{r}=0,0,AVERAGEIFS({m},{b},{c},{g},"F"))'
                  '+IF(C{r}=0,0,SUMIFS({z},{b},{c},{g},"F")/C{r}))').format(
            r=r, m=M_, b=B, c=cf, g=G_, z="%s$Z$%d:$Z$%d" % (D, first, last))
        put(41, "=IF(E{r}<=0,0,(E{r}-{nf})/E{r}*100-X{r})".format(r=r, nf=new_f2), PCT)

        put(42, "=IF(OR(B{r}<10,C{r}<3,D{r}<3),1,0)".format(r=r))
        put(43, "=IF(H{r}<0,1,0)".format(r=r))

    clast = CHDR + len(order)

    # ======================================================================
    # ЛИСТ Summary
    # ======================================================================
    sm = wb.create_sheet("Summary")
    sm["A1"] = "Totals"
    sm["A1"].font = TITLE
    sm["A2"] = ("Every figure here is a formula referencing the Data and Categories sheets. "
                "No number on this sheet is typed in by hand.")
    sm["A2"].font = MUTED
    sm.column_dimensions["A"].width = 40
    for col in "BCD":
        sm.column_dimensions[col].width = 20

    sm["A4"] = "Payroll before adjustment"
    sm["A4"].font = BOLD
    sm["B4"] = "=SUM(Data!$N${f}:$N${l})".format(f=first, l=last)
    sm["B4"].number_format = EUR
    sm["B4"].fill = CALC_FILL

    sm["A5"] = "Headcount"
    sm["B5"] = "=COUNTA(Data!$A${f}:$A${l})".format(f=first, l=last)
    sm["B5"].fill = CALC_FILL

    SHDR = 7
    header(sm, SHDR, ["Metric", "Minimum compliance", "Full equalisation", "Difference"])
    # номера строк вычисляются от SHDR, а не вписываются константами:
    # раньше они разъехались с фактической разметкой и «Итого» ссылалось само
    # на себя, давая циклическую ссылку и пустую ячейку
    R_ADJ = SHDR + 1        # корректировка ФОТ
    R_CON = SHDR + 2        # взносы
    R_TOT = SHDR + 3        # итого
    TOTAL_LABEL = "Total"
    lines = [
        ("Pay adjustment", "=SUM(Categories!$AC${f}:$AC${l})",
         "=SUM(Categories!$AK${f}:$AK${l})", EUR),
        ("Employer contributions", "=SUM(Categories!$AD${f}:$AD${l})",
         "=SUM(Categories!$AL${f}:$AL${l})", EUR),
        (TOTAL_LABEL, "=B{a}+B{c}".format(a=R_ADJ, c=R_CON),
         "=C{a}+C{c}".format(a=R_ADJ, c=R_CON), EUR),
        ("As % of payroll", "=IF($B$4=0,0,B{t}/$B$4*100)".format(t=R_TOT),
         "=IF($B$4=0,0,C{t}/$B$4*100)".format(t=R_TOT), PCT),
        ("Employees receiving an uplift", "=SUM(Categories!$AF${f}:$AF${l})",
         "=SUM(Categories!$AN${f}:$AN${l})", '#,##0'),
        ("Current-year effect",
         "=B{t}*(13-implementation_month)/12".format(t=R_TOT),
         "=C{t}*(13-implementation_month)/12".format(t=R_TOT), EUR),
        ("Next-year effect", "=B{t}".format(t=R_TOT),
         "=C{t}".format(t=R_TOT), EUR),
    ]
    for i, (label, fmin, ffull, fmt) in enumerate(lines):
        r = SHDR + 1 + i
        c = sm.cell(row=r, column=1, value=label)
        c.border = BOX
        if label == TOTAL_LABEL:
            c.font = BOLD
        for col, f in ((2, fmin), (3, ffull)):
            cc = sm.cell(row=r, column=col, value=f.format(f=CHDR + 1, l=clast))
            cc.number_format = fmt
            cc.fill = CALC_FILL
            cc.border = BOX
            if label == TOTAL_LABEL:
                cc.font = BOLD
        d = sm.cell(row=r, column=4, value="=C{r}-B{r}".format(r=r))
        d.number_format = fmt
        d.fill = CALC_FILL
        d.border = BOX
        if label == TOTAL_LABEL:
            d.font = BOLD

    # строки ниже таблицы — от её фактического конца, не константами
    r_note = SHDR + len(lines) + 2
    sm.cell(row=r_note, column=1, value="Reconciliation with the calculator").font = BOLD
    note_cell = sm.cell(row=r_note + 1, column=1, value=(
        "Filled in at the reconciliation step: totals from the HTML calculator are entered "
        "here and compared against B{t} / C{t}. Any discrepancy above one cent is a blocker."
        .format(t=R_TOT)))
    note_cell.font = MUTED
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    sm.merge_cells(start_row=r_note + 1, start_column=1,
                   end_row=r_note + 2, end_column=4)

    # ======================================================================
    # ЛИСТ Legend
    # ======================================================================
    lg = wb.create_sheet("Legend")
    lg["A1"] = "Legend and method"
    lg["A1"].font = TITLE
    lg.column_dimensions["A"].width = 30
    lg.column_dimensions["B"].width = 110

    blocks = [
        ("Colours", ""),
        ("Yellow", "Source data. Comes from the CSV, never calculated."),
        ("Blue", "Calculated by formula. Nothing to edit — it recalculates itself."),
        ("Green", "A setting on the Settings sheet. These are meant to be changed."),
        ("", ""),
        ("Units", ""),
        ("Where to find them",
         "The units for each column are in a comment on its header cell on the Data sheet "
         "(red corner marker — hover over it). The same information, in short, is below."),
        ("base_salary, variable_pay",
         "Euros actually paid for the period, not annualised. Someone on a 0.5 FTE contract, "
         "or who worked 7 months, shows what they actually received. Variable pay follows the "
         "same convention as base salary."),
        ("n_base, n_var, n_total",
         "Euros per year at full time, obtained by dividing actual pay by factor. Employees can "
         "only be compared on these figures: otherwise half the gap would be a difference in "
         "working time rather than in pay."),
        ("grade",
         "Position in the grade ladder: a higher number means a higher grade. Only the ordering "
         "matters to the calculation; the absolute value carries no meaning."),
        ("", ""),
        ("Order of calculation", ""),
        ("1. Normalisation",
         "factor = fte x (months_worked / 12). Base salary and variable pay are divided by factor, "
         "converting them to a full-time, full-year equivalent. Only normalised figures are "
         "comparable across employees."),
        ("2. Gaps",
         "raw_gap = (mean_M - mean_F) / mean_M x 100, computed on n_total (normalised base plus "
         "variable pay), separately for each category."),
        ("3. Regression",
         "OLS of n_total on grade and tenure within each category, solved from explicit sums via "
         "Cramer's rule. It answers one question: how much of the gap is accounted for by grade "
         "and tenure. The guard mirrors calc.js — fewer than 4 observations or a constant "
         "predictor means the regression is undefined, explained = 0, and the whole gap is "
         "treated as unexplained (the conservative direction)."),
        ("4. Explained share",
         "explained = coef_grade x (mean male grade - mean female grade) + coef_tenure x "
         "(mean male tenure - mean female tenure), expressed as % of mean_M. Capped at the size "
         "of the gap itself: the model cannot explain more than exists."),
        ("5. Selecting recipients",
         "Women whose normalised n_total falls below the median of their own category (the median "
         "across the whole category, both genders). Categories where the gap favours women are "
         "not adjusted in either scenario — levelling only ever moves pay upwards."),
        ("6. Distribution",
         "Proportional to the shortfall against the median: share = (median - n_total) / sum of "
         "all shortfalls. The further below the middle an employee sits, the larger the portion "
         "of the budget they receive."),
        ("Two distinct quantities",
         "The average category gap and an individual's shortfall do different jobs. The average gap "
         "determines HOW MUCH MONEY the category needs in total (step 5, the need formula). The "
         "shortfall against the median determines HOW THAT MONEY IS SPLIT between people (step 6). "
         "The average gap does not enter an individual uplift: a person's shortfall is measured "
         "against the category median, not against the male average."),
        ("Consequence",
         "A woman paid above the median receives nothing, even where her category's gap is large. "
         "The method lifts the bottom half rather than every woman across the board."),
        ("7. Contributions",
         "Charged on the increase in base pay, not on total salary, and always against actual base "
         "pay. The portion of the uplift that fits under the ceiling attracts the full rate; the "
         "portion above it attracts the reduced rate."),
        ("8. Totals",
         "Current-year effect = full amount x (13 - implementation month) / 12. Implementing in "
         "January gives 12/12, in December 1/12."),
        ("", ""),
        ("The two scenarios", ""),
        ("Minimum compliance",
         "Bring the unexplained gap down to the threshold (5% by default). Categories already "
         "below the threshold are left untouched."),
        ("Full equalisation",
         "Bring the unexplained gap to zero in every category where it is positive."),
        ("", ""),
        ("Caveats", ""),
        ("What explained means",
         "The word is technical, not exculpatory. If women are systematically not promoted through "
         "the grades, the regression will attribute the gap to grade and call it explained. That is "
         "a limitation of the method, not a finding about fairness."),
        ("No array formulas",
         "The whole model is built on ordinary formulas: SUMIFS, SUMPRODUCT, COUNTIFS. "
         "Ctrl+Shift+Enter is never required, in any version of Excel. Medians are computed via "
         "rank (the helper columns _rank_cat / _rank_sex on the Data sheet) and the regression "
         "from explicit sums rather than LINEST."),
        ("PRNG",
         "The demo data is generated by mulberry32 + Box-Muller from a fixed seed. That cannot be "
         "reproduced byte-for-byte in Excel formulas, so the array is exported as plain numbers. "
         "Everything computed on top of it is live."),
    ]
    r = 3
    r = 3
    for label, text in blocks:
        a = lg.cell(row=r, column=1, value=label)
        b = lg.cell(row=r, column=2, value=text)
        if text == "":
            a.font = BOLD
        else:
            a.font = Font(bold=True, size=10)
            b.alignment = Alignment(wrap_text=True, vertical="top")
            lg.row_dimensions[r].height = max(15, 13 * (len(text) // 105 + 1))
        r += 1

    wb.save(out_path)
    return n, len(order)


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "build/sample.csv"
    dst = sys.argv[2] if len(sys.argv) > 2 else "build/excel/pay-gap-model-mini.xlsx"
    note = sys.argv[3] if len(sys.argv) > 3 else "Мини-версия на демо-фикстуре (10 строк)"
    rows, cats = build(src, dst, note)
    print("OK: %s  (%d strok, %d kategoriy)" % (dst, rows, cats))
